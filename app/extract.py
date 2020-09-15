import sys
import json
from pathlib import Path
from collections import namedtuple
from operator import itemgetter

from openpyxl import load_workbook

from app import db
from app.models import Country, Basket, PurchasingPowerParity, ExchangeRate


StyleSet = namedtuple("StyleSet", ["font_size", "bold", "fill_color"])
ExchangeRateRow = namedtuple("ExchangeRateRow", ["indicator_code", "country_name", "period", "currency_code", "value", "xc"])
Year = namedtuple("Year", ["basket_names", "country_codes", "data"])

TOTAL_COLUMN_NAME = "WAS"
EXCHANGE_RATE_COLUMN_NAMES = ["IndicatorCode", "CountryName", "Period", "CurrencyCode", "Value", "XC"]
EXCHANGE_RATE_INDICATOR_CODE = "EXCHANGE RATE"
XC_TEMPLATE = "{}/USD"


def extract_data(ppp_spreadsheet_filepath, exchange_rate_spreadsheet_filepath, country_code_map_json_filepath):
    ppp_spreadsheet_filepath = Path(ppp_spreadsheet_filepath)
    exchange_rate_spreadsheet_filepath = Path(exchange_rate_spreadsheet_filepath)
    country_code_map_json_filepath = Path(country_code_map_json_filepath)

    assert ppp_spreadsheet_filepath.exists()
    assert exchange_rate_spreadsheet_filepath.exists()
    assert country_code_map_json_filepath.exists()

    ppp_wb = load_workbook(filename=ppp_spreadsheet_filepath)
    num_years = len(ppp_wb.sheetnames)
    years_ppps_dict  = {}
    for year_idx in range(num_years):
        sheetname = ppp_wb.sheetnames[year_idx]
        year = int(sheetname)
        ws = ppp_wb[sheetname]

        year_basket_names = []
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            basket_name = row[0].value
            if basket_name is None:
                break
            year_basket_names.append(basket_name)

        col_cells = [col[0] for col in ws.iter_cols(min_col=3, min_row=1, max_row=1)]
        year_country_codes = [col_cells[i].value for i in range(len(col_cells) - 1)]
        year_data = [None] * len(year_basket_names) * len(year_country_codes)

        row_idx = 0
        for row in ws.iter_rows(min_row=2, min_col=3, max_row=len(year_basket_names) + 1, max_col=len(year_country_codes) + 2):
            offset = len(year_country_codes) * row_idx
            year_data[offset : offset + len(year_country_codes)] = [cell.value for cell in row]
            row_idx += 1

        years_ppps_dict[year] = Year(year_basket_names, year_country_codes, year_data)

    with open(country_code_map_json_filepath) as country_code_map_json_file:
        country_code_map = json.load(country_code_map_json_file)

    country_name_map = {country_code_map[country_code][0]: (country_code, country_code_map[country_code][1]) for country_code in country_code_map}

    exchange_rate_wb = load_workbook(filename=exchange_rate_spreadsheet_filepath)
    exchange_rate_ws = exchange_rate_wb["Sheet1"]
    column_names_row = [row for row in exchange_rate_ws.iter_rows(min_row=1, max_row=1, min_col=1)][0]
    column_name_cells = []
    for cell in column_names_row:
        value = cell.value
        if value is None:
            break
        column_name_cells.append(value)
    assert column_name_cells == EXCHANGE_RATE_COLUMN_NAMES

    years_exchange_rates_dict = {}
    for row in exchange_rate_ws.iter_rows(min_row=2, min_col=1, max_col=len(EXCHANGE_RATE_COLUMN_NAMES)):
        exchange_rate_row = ExchangeRateRow(*[cell.value for cell in row])
        if exchange_rate_row.indicator_code is None:
            break
        assert exchange_rate_row.indicator_code == EXCHANGE_RATE_INDICATOR_CODE
        country_code, currency_code = country_name_map[exchange_rate_row.country_name]
        year = int(exchange_rate_row.period)
        assert exchange_rate_row.currency_code == currency_code
        exchange_rate = float(exchange_rate_row.value)
        assert exchange_rate_row.xc == XC_TEMPLATE.format(exchange_rate_row.currency_code)
        try:
            year_exchange_rates = years_exchange_rates_dict[year]
        except KeyError:
            year_exchange_rates = years_exchange_rates_dict[year] = {}
        assert country_code not in year_exchange_rates
        year_exchange_rates[country_code] = exchange_rate

    unique_country_codes = set()
    unique_basket_names = set()
    for year in years_ppps_dict:
        year_ppps = years_ppps_dict[year]
        unique_country_codes.update(year_ppps.country_codes)
        unique_basket_names.update(year_ppps.basket_names)

    basket_names_sorted = sorted(unique_basket_names)

    PurchasingPowerParity.query.delete()
    ExchangeRate.query.delete()
    Basket.query.delete()
    Country.query.delete()

    for code in sorted(unique_country_codes):
        name, currency_code = country_code_map[code]
        country = Country(name=name, currency_code=currency_code)
        db.session.add(country)

    for name in basket_names_sorted:
        basket = Basket(name=name)
        db.session.add(basket)

    for year in years_ppps_dict:
        year_obj = years_ppps_dict[year]

        countries = [Country.query.filter_by(currency_code=country_code_map[country_code][1]).first() for country_code in year_obj.country_codes]
        baskets = [Basket.query.filter_by(name=basket_name).first() for basket_name in year_obj.basket_names]
        data = year_obj.data
        i = 0
        for basket in baskets:
            for country in countries:
                ppp = PurchasingPowerParity(year=year, country=country, basket=basket, value=data[i])
                db.session.add(ppp)
                i += 1

    for year in years_exchange_rates_dict:
        year_exchange_rates_dict = years_exchange_rates_dict[year]
        for country_code in year_exchange_rates_dict:
            value = year_exchange_rates_dict[country_code]
            country_name, currency_code = country_code_map[country_code]
            country = Country.query.filter_by(currency_code=currency_code).first()
            exchange_rate = ExchangeRate(year=year, country=country, value=value)
            db.session.add(exchange_rate)
